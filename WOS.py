"""
Work Ledger Pro - Streamlit Version
400kV Ballari Pooling Station

Converted from the Flet Card UI version to Streamlit.
Default login:
    User ID  : admin
    Password : admin@123

Run:
    streamlit run work_ledger_streamlit.py
"""

import json
import os
import shutil
import uuid
from datetime import date, datetime
from io import BytesIO
from pathlib import Path

import pandas as pd
import streamlit as st

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except Exception:
    OPENPYXL_AVAILABLE = False

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    REPORTLAB_AVAILABLE = True
except Exception:
    REPORTLAB_AVAILABLE = False

# ============================================================
# CONFIGURATION
# ============================================================
APP_TITLE = "Work Ledger Pro - 400kV Ballari PS"
APP_HEADER = "400kV BPS WORK LEDGER PRO"
FOOTER_TEXT = "Created by Manjunath Raju P"
LOGIN_USER = "admin"
LOGIN_PASSWORD = "admin@123"

USER_HOME = os.path.expanduser("~")
ONEDRIVE_DIR = os.path.join(USER_HOME, "OneDrive", "Work_Ledger_App")
UPLOAD_DIR = os.path.join(ONEDRIVE_DIR, "work_docs")
BACKUP_DIR = os.path.join(ONEDRIVE_DIR, "backups")
DATA_FILE = os.path.join(ONEDRIVE_DIR, "ledger_data.json")
CONFIG_FILE = os.path.join(ONEDRIVE_DIR, "ledger_config.json")
MAX_BACKUPS = 50

DOC_TYPES = [
    "Report", "Request", "Enquiry", "Quotation", "Comparative", "Estimate",
    "PO/Work Award", "Bill", "CR", "MOM", "Remarks"
]
FY_OPTIONS = ["2025-26", "2026-27", "2027-28"]
ACCOUNT_HEADS = ["74.110", "74.510", "74.116", "76.220", "76.153"]
STATION_OPTIONS = ["400kV Ballari Pooling Station"]

Path(UPLOAD_DIR).mkdir(parents=True, exist_ok=True)
Path(BACKUP_DIR).mkdir(parents=True, exist_ok=True)

# ============================================================
# HELPERS
# ============================================================
def clean_amount(value) -> float:
    if value is None:
        return 0.0
    try:
        sanitized = str(value).replace(",", "").replace("₹", "").strip()
        return float(sanitized) if sanitized else 0.0
    except Exception:
        return 0.0


def money(value) -> str:
    return f"₹{clean_amount(value):,.2f}"


def normalize_wo_no(value) -> str:
    return str(value or "").strip().lower()


def parse_date_string(value):
    raw = str(value or "").strip()
    if not raw:
        return None
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(raw, fmt).date().isoformat()
        except ValueError:
            pass
    return None


def normalize_files_dict(files):
    result = {doc_type: [] for doc_type in DOC_TYPES}
    if not isinstance(files, dict):
        return result
    for doc_type in DOC_TYPES:
        val = files.get(doc_type)
        if val is None or val == "":
            result[doc_type] = []
        elif isinstance(val, list):
            result[doc_type] = [str(x) for x in val if x]
        else:
            result[doc_type] = [str(val)]
    return result


def atomic_write_json(path, payload):
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    tmp_path = path + ".tmp"
    with open(tmp_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=4, ensure_ascii=False)
    os.replace(tmp_path, path)


def cleanup_old_backups():
    files = sorted(Path(BACKUP_DIR).glob("backup_*.json"), key=lambda p: p.stat().st_mtime, reverse=True)
    for old_file in files[MAX_BACKUPS:]:
        try:
            old_file.unlink()
        except Exception:
            pass


def save_data(data):
    try:
        atomic_write_json(DATA_FILE, data)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        shutil.copy2(DATA_FILE, os.path.join(BACKUP_DIR, f"backup_{ts}.json"))
        cleanup_old_backups()
        return True, "Data saved successfully."
    except Exception as e:
        return False, f"Error saving data: {e}"


def load_data():
    if not os.path.exists(DATA_FILE):
        return []
    try:
        with open(DATA_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
        if not isinstance(data, list):
            return []
        for item in data:
            if "id" not in item:
                item["id"] = str(uuid.uuid4())
            item["Files"] = normalize_files_dict(item.get("Files", {}))
            item["Amount"] = clean_amount(item.get("Amount", 0))
        return data
    except Exception:
        return []


def load_config():
    if not os.path.exists(CONFIG_FILE):
        return {"budget": 553000.0}
    try:
        with open(CONFIG_FILE, "r", encoding="utf-8") as f:
            cfg = json.load(f)
        if not isinstance(cfg, dict):
            return {"budget": 553000.0}
        return cfg
    except Exception:
        return {"budget": 553000.0}


def save_config(config):
    try:
        atomic_write_json(CONFIG_FILE, config)
        return True, "Budget saved."
    except Exception as e:
        return False, f"Budget save failed: {e}"


def safe_rerun():
    try:
        st.rerun()
    except Exception:
        st.experimental_rerun()


def make_dataframe(records):
    rows = []
    running = 0.0
    for i, item in enumerate(records, start=1):
        amt = clean_amount(item.get("Amount"))
        running += amt
        rows.append({
            "Sl": i,
            "FY": item.get("FY", ""),
            "Station": item.get("Station Name", ""),
            "Work Name": item.get("Work Name", ""),
            "Account Head": item.get("Account Head", ""),
            "WO No": item.get("WO No", ""),
            "WO Date": item.get("WO Date", ""),
            "Amount": amt,
            "Cumulative Total": running,
        })
    return pd.DataFrame(rows)


def filter_records(data, fy, station, account_text, wo_text, work_text):
    result = []
    for item in data:
        if fy != "All" and item.get("FY") != fy:
            continue
        if station != "All" and item.get("Station Name") != station:
            continue
        if account_text and account_text.strip() not in str(item.get("Account Head", "")):
            continue
        if wo_text and wo_text.strip().lower() not in str(item.get("WO No", "")).lower():
            continue
        if work_text and work_text.strip().lower() not in str(item.get("Work Name", "")).lower():
            continue
        result.append(item)
    return result


def validate_entry(work_name, wo_no, wo_date_raw, amount, data, editing_id=None):
    work_name = str(work_name or "").strip()
    wo_no = str(wo_no or "").strip()
    wo_date = parse_date_string(wo_date_raw)
    amount = clean_amount(amount)

    if not work_name:
        return False, "Please enter Work Name.", None
    if not wo_no:
        return False, "Please enter WO No.", None
    if not wo_date:
        return False, "Invalid WO Date. Use YYYY-MM-DD or DD-MM-YYYY.", None
    if amount <= 0:
        return False, "Please enter valid Amount.", None

    norm_wo = normalize_wo_no(wo_no)
    for item in data:
        if normalize_wo_no(item.get("WO No")) == norm_wo and item.get("id") != editing_id:
            return False, "Duplicate WO No found. Please check the existing entry.", None

    return True, "OK", {
        "Work Name": work_name,
        "WO No": wo_no,
        "WO Date": wo_date,
        "Amount": amount,
    }


def build_excel(records):
    if not OPENPYXL_AVAILABLE:
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = "Ledger"
    headers = ["Sl", "FY", "Station", "Work Name", "Account Head", "WO No", "WO Date", "Amount", "Cumulative Total"]
    ws.append(headers)

    running = 0.0
    for i, item in enumerate(records, start=1):
        amt = clean_amount(item.get("Amount"))
        running += amt
        ws.append([
            i,
            item.get("FY"),
            item.get("Station Name"),
            item.get("Work Name"),
            item.get("Account Head"),
            item.get("WO No"),
            item.get("WO Date"),
            amt,
            running,
        ])
    ws.append(["", "", "", "", "", "", "TOTAL", running, ""])

    header_fill = PatternFill("solid", fgColor="1565C0")
    white_font = Font(color="FFFFFF", bold=True)
    total_fill = PatternFill("solid", fgColor="E3F2FD")
    thin = Side(style="thin", color="D6DEE8")

    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = white_font

    for cell in ws[ws.max_row]:
        cell.fill = total_fill
        cell.font = Font(bold=True)

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 55
    ws.column_dimensions["H"].number_format = '₹#,##0.00'
    ws.column_dimensions["I"].number_format = '₹#,##0.00'

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()


def build_pdf(records):
    if not REPORTLAB_AVAILABLE:
        return None

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=18,
        leftMargin=18,
        topMargin=18,
        bottomMargin=18,
    )
    styles = getSampleStyleSheet()
    elements = [
        Paragraph("WORK LEDGER REPORT - 400kV Ballari PS", styles["Title"]),
        Spacer(1, 10),
    ]

    headers = ["Sl", "FY", "Station", "Work Name", "Head", "WO No", "Date", "Amount"]
    rows = [headers]
    total = 0.0
    for i, item in enumerate(records, start=1):
        amt = clean_amount(item.get("Amount"))
        total += amt
        rows.append([
            i,
            item.get("FY", ""),
            item.get("Station Name", ""),
            str(item.get("Work Name", ""))[:55],
            item.get("Account Head", ""),
            item.get("WO No", ""),
            item.get("WO Date", ""),
            f"{amt:,.2f}",
        ])
    rows.append(["", "", "", "", "", "", "TOTAL", f"{total:,.2f}"])

    table = Table(rows, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1565C0")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#E3F2FD")),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
    ]))
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()


def save_uploaded_attachment(item, doc_type, uploaded_file):
    if uploaded_file is None:
        return False, "Please choose a file first."
    try:
        original_name = uploaded_file.name
        safe_name = os.path.basename(original_name)
        stored_name = f"{date.today()}_{uuid.uuid4().hex[:6]}_{safe_name}"
        target = os.path.join(UPLOAD_DIR, stored_name)
        with open(target, "wb") as f:
            f.write(uploaded_file.getbuffer())

        item["Files"] = normalize_files_dict(item.get("Files", {}))
        item["Files"].setdefault(doc_type, [])
        item["Files"][doc_type].append(stored_name)
        ok, msg = save_data(st.session_state.data)
        return ok, "File uploaded successfully." if ok else msg
    except Exception as e:
        return False, f"Upload failed: {e}"


# ============================================================
# STREAMLIT UI
# ============================================================
st.set_page_config(page_title=APP_TITLE, page_icon="📒", layout="wide")

st.markdown(
    """
    <style>
    .main { background-color: #F4F7FB; }
    .block-container { padding-top: 1.2rem; padding-bottom: 2rem; }
    .app-header {
        background: #0B1F3A;
        color: white;
        padding: 18px 24px;
        border-radius: 18px;
        margin-bottom: 18px;
    }
    .app-header h1 { margin: 0; font-size: 28px; }
    .app-header p { margin: 3px 0 0 0; color: #DDEBFF; }
    .section-card {
        background: #FFFFFF;
        border: 1px solid #D6DEE8;
        border-radius: 16px;
        padding: 18px;
        margin-bottom: 16px;
    }
    .small-muted { color: #607080; font-size: 12px; }
    .footer { color: #607080; font-style: italic; text-align: right; font-size: 12px; }
    </style>
    """,
    unsafe_allow_html=True,
)

if "logged_in" not in st.session_state:
    st.session_state.logged_in = False
if "data" not in st.session_state:
    st.session_state.data = load_data()
if "budget" not in st.session_state:
    st.session_state.budget = clean_amount(load_config().get("budget", 553000.0))
if "delete_confirm" not in st.session_state:
    st.session_state.delete_confirm = False


def refresh_from_disk():
    st.session_state.data = load_data()
    st.session_state.budget = clean_amount(load_config().get("budget", 553000.0))


# -------------------------
# LOGIN SCREEN
# -------------------------
if not st.session_state.logged_in:
    st.markdown("<div class='app-header'><h1>🔐 Work Ledger Login</h1><p>400kV Ballari Pooling Station</p></div>", unsafe_allow_html=True)

    login_col, _ = st.columns([1, 2])
    with login_col:
        with st.form("login_form"):
            user_id = st.text_input("User ID", value="admin")
            password = st.text_input("Password", type="password")
            submitted = st.form_submit_button("Login", use_container_width=True)

        st.caption("Default: admin / admin@123")
        if submitted:
            if user_id.strip() == LOGIN_USER and password.strip() == LOGIN_PASSWORD:
                st.session_state.logged_in = True
                safe_rerun()
            else:
                st.error("Invalid User ID or Password")
    st.stop()


# -------------------------
# MAIN SCREEN
# -------------------------
header_left, header_right = st.columns([5, 1])
with header_left:
    st.markdown(
        f"<div class='app-header'><h1>{APP_HEADER}</h1><p>FY 2026-2027</p></div>",
        unsafe_allow_html=True,
    )
with header_right:
    st.write("")
    st.write("")
    if st.button("Logout", use_container_width=True):
        st.session_state.logged_in = False
        safe_rerun()
    if st.button("Reload Data", use_container_width=True):
        refresh_from_disk()
        st.success("Data reloaded from disk.")

# Sidebar filters
with st.sidebar:
    st.header("🔍 Filters")
    all_stations = sorted(set(STATION_OPTIONS + [d.get("Station Name", "") for d in st.session_state.data if d.get("Station Name")]))
    f_fy = st.selectbox("FY", ["All"] + FY_OPTIONS, index=0, key="filter_fy")
    f_station = st.selectbox("Station", ["All"] + all_stations, index=0, key="filter_station")
    f_account = st.text_input("Account Head contains", key="filter_account")
    f_wo = st.text_input("WO No contains", key="filter_wo")
    f_work = st.text_input("Work Name contains", key="filter_work")
    st.divider()
    st.caption(f"Data path: {DATA_FILE}")
    st.caption(f"Attachment path: {UPLOAD_DIR}")

filtered_data = filter_records(st.session_state.data, f_fy, f_station, f_account, f_wo, f_work)

total_spent = sum(clean_amount(d.get("Amount")) for d in st.session_state.data)
filtered_total = sum(clean_amount(d.get("Amount")) for d in filtered_data)
available = st.session_state.budget - total_spent

m1, m2, m3, m4 = st.columns(4)
m1.metric("Budget", money(st.session_state.budget))
m2.metric("Total Spent", money(total_spent))
m3.metric("Filtered Total", money(filtered_total))
m4.metric("Available", money(available))

# Budget control
with st.container(border=True):
    st.subheader("⚙️ Budget Control")
    bcol1, bcol2 = st.columns([1, 5])
    with bcol1:
        new_budget = st.number_input("Budget (₹)", min_value=0.0, value=float(st.session_state.budget), step=1000.0, format="%.2f")
        if st.button("Save Budget", use_container_width=True):
            if new_budget <= 0:
                st.error("Enter valid budget amount.")
            else:
                st.session_state.budget = clean_amount(new_budget)
                ok, msg = save_config({"budget": st.session_state.budget})
                st.success(msg) if ok else st.error(msg)
    with bcol2:
        st.info(f"Showing {len(filtered_data)} of {len(st.session_state.data)} entries")

# Export buttons
with st.container(border=True):
    st.subheader("📤 Export Filtered Report")
    ecol1, ecol2, ecol3 = st.columns([1, 1, 3])
    report_name_date = date.today().isoformat()

    excel_bytes = build_excel(filtered_data)
    if excel_bytes:
        ecol1.download_button(
            "Download Excel",
            data=excel_bytes,
            file_name=f"Ledger_Report_{report_name_date}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    else:
        ecol1.warning("Install openpyxl for Excel export.")

    pdf_bytes = build_pdf(filtered_data)
    if pdf_bytes:
        ecol2.download_button(
            "Download PDF",
            data=pdf_bytes,
            file_name=f"Ledger_Report_{report_name_date}.pdf",
            mime="application/pdf",
            use_container_width=True,
        )
    else:
        ecol2.warning("Install reportlab for PDF export.")

# Ledger table
with st.container(border=True):
    st.subheader("📊 Ledger List")
    ledger_df = make_dataframe(filtered_data)
    if ledger_df.empty:
        st.info("No ledger entries found for the selected filter.")
    else:
        formatted_df = ledger_df.copy()
        formatted_df["Amount"] = formatted_df["Amount"].map(money)
        formatted_df["Cumulative Total"] = formatted_df["Cumulative Total"].map(money)
        st.dataframe(formatted_df, hide_index=True, use_container_width=True, height=420)

# Add entry
with st.container(border=True):
    st.subheader("➕ Add New Entry")
    with st.form("add_entry_form", clear_on_submit=True):
        a1, a2, a3 = st.columns([1, 2, 1])
        fy = a1.selectbox("FY", FY_OPTIONS, key="add_fy")
        station = a2.selectbox("Station", STATION_OPTIONS, key="add_station")
        account = a3.selectbox("Account Head", ACCOUNT_HEADS, key="add_account")

        work_name = st.text_area("Work Name", height=80, key="add_work")
        c1, c2, c3 = st.columns(3)
        wo_no = c1.text_input("WO No", key="add_wo")
        wo_date = c2.text_input("WO Date", value=str(date.today()), help="Accepted: YYYY-MM-DD / DD-MM-YYYY / DD/MM/YYYY", key="add_date")
        amount = c3.number_input("Amount (₹)", min_value=0.0, step=1000.0, format="%.2f", key="add_amount")

        add_submitted = st.form_submit_button("Save Entry", use_container_width=True)

    if add_submitted:
        ok, msg, validated = validate_entry(work_name, wo_no, wo_date, amount, st.session_state.data)
        if not ok:
            st.error(msg)
        else:
            new_item = {
                "id": str(uuid.uuid4()),
                "FY": fy,
                "Station Name": station,
                "Account Head": account,
                **validated,
                "Files": {doc_type: [] for doc_type in DOC_TYPES},
            }
            st.session_state.data.append(new_item)
            saved, save_msg = save_data(st.session_state.data)
            if saved:
                st.success("Entry saved successfully.")
                safe_rerun()
            else:
                st.error(save_msg)

# Edit/Delete/Attachments
with st.container(border=True):
    st.subheader("✏️ Edit / Delete / Attach Documents")
    if not st.session_state.data:
        st.info("No entries available for editing.")
    else:
        choices = []
        id_map = {}
        for item in st.session_state.data:
            label = f"{item.get('WO No', '')} | {item.get('Work Name', '')[:80]}"
            choices.append(label)
            id_map[label] = item.get("id")

        selected_label = st.selectbox("Select Entry", choices)
        selected_id = id_map[selected_label]
        selected_item = next((d for d in st.session_state.data if d.get("id") == selected_id), None)

        if selected_item:
            with st.form("edit_entry_form"):
                e1, e2, e3 = st.columns([1, 2, 1])
                edit_fy = e1.selectbox("FY", FY_OPTIONS, index=FY_OPTIONS.index(selected_item.get("FY", FY_OPTIONS[0])) if selected_item.get("FY") in FY_OPTIONS else 0, key=f"edit_fy_{selected_id}")
                edit_station_options = sorted(set(STATION_OPTIONS + [selected_item.get("Station Name", "")]))
                edit_station = e2.selectbox("Station", edit_station_options, index=edit_station_options.index(selected_item.get("Station Name", edit_station_options[0])) if selected_item.get("Station Name") in edit_station_options else 0, key=f"edit_station_{selected_id}")
                edit_account = e3.selectbox("Account Head", ACCOUNT_HEADS, index=ACCOUNT_HEADS.index(selected_item.get("Account Head", ACCOUNT_HEADS[0])) if selected_item.get("Account Head") in ACCOUNT_HEADS else 0, key=f"edit_account_{selected_id}")

                edit_work = st.text_area("Work Name", value=selected_item.get("Work Name", ""), height=80, key=f"edit_work_{selected_id}")
                c1, c2, c3 = st.columns(3)
                edit_wo = c1.text_input("WO No", value=selected_item.get("WO No", ""), key=f"edit_wo_{selected_id}")
                edit_date = c2.text_input("WO Date", value=selected_item.get("WO Date", str(date.today())), key=f"edit_date_{selected_id}")
                edit_amount = c3.number_input("Amount (₹)", min_value=0.0, value=float(clean_amount(selected_item.get("Amount"))), step=1000.0, format="%.2f", key=f"edit_amount_{selected_id}")

                update_clicked = st.form_submit_button("Update Entry", use_container_width=True)

            if update_clicked:
                ok, msg, validated = validate_entry(edit_work, edit_wo, edit_date, edit_amount, st.session_state.data, editing_id=selected_id)
                if not ok:
                    st.error(msg)
                else:
                    selected_item.update({
                        "FY": edit_fy,
                        "Station Name": edit_station,
                        "Account Head": edit_account,
                        **validated,
                    })
                    selected_item["Files"] = normalize_files_dict(selected_item.get("Files", {}))
                    saved, save_msg = save_data(st.session_state.data)
                    if saved:
                        st.success("Entry updated successfully.")
                        safe_rerun()
                    else:
                        st.error(save_msg)

            d1, d2 = st.columns([1, 5])
            with d1:
                delete_check = st.checkbox("Confirm delete", key=f"delete_check_{selected_id}")
                if st.button("Delete Entry", type="primary", disabled=not delete_check, use_container_width=True):
                    st.session_state.data = [d for d in st.session_state.data if d.get("id") != selected_id]
                    saved, save_msg = save_data(st.session_state.data)
                    if saved:
                        st.success("Entry deleted.")
                        safe_rerun()
                    else:
                        st.error(save_msg)
            with d2:
                st.warning("Delete removes the ledger entry. Existing uploaded files remain in the storage folder.")

            st.divider()
            st.subheader("📂 Upload Attachment")
            u1, u2 = st.columns([1, 3])
            doc_type = u1.selectbox("Document Type", DOC_TYPES, key=f"doc_type_{selected_id}")
            uploaded_file = u2.file_uploader("Choose file to attach to selected entry", type=None, key=f"upload_file_{selected_id}")
            if st.button("Upload File", use_container_width=True, key=f"upload_btn_{selected_id}"):
                ok, msg = save_uploaded_attachment(selected_item, doc_type, uploaded_file)
                st.success(msg) if ok else st.error(msg)
                if ok:
                    safe_rerun()

            st.subheader("📥 Attached Documents")
            selected_item["Files"] = normalize_files_dict(selected_item.get("Files", {}))
            any_file = False
            for doc_type_name, filenames in selected_item["Files"].items():
                for filename in list(filenames):
                    any_file = True
                    file_path = os.path.join(UPLOAD_DIR, filename)
                    row1, row2, row3, row4 = st.columns([1, 4, 1, 1])
                    row1.markdown(f"**{doc_type_name}**")
                    row2.write(filename)
                    if os.path.exists(file_path):
                        with open(file_path, "rb") as f:
                            row3.download_button("Download", data=f.read(), file_name=filename, key=f"download_{doc_type_name}_{filename}")
                    else:
                        row3.error("Missing")
                    if row4.button("Remove", key=f"remove_{doc_type_name}_{filename}"):
                        selected_item["Files"][doc_type_name].remove(filename)
                        saved, save_msg = save_data(st.session_state.data)
                        st.success("Attachment removed from record." if saved else save_msg)
                        safe_rerun()
            if not any_file:
                st.info("No attachments available for this entry.")

st.markdown(f"<div class='footer'>{FOOTER_TEXT}</div>", unsafe_allow_html=True)
